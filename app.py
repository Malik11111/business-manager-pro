"""
Business Manager Pro — Flask Backend
Sert l'application web + API REST + authentification
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from config import Config
from models import db, User, Etablissement, Prestataire, Evaluation, CorbeillePresta
from models import Personnel, Unite, Materiel
from models import PharmaStock, PharmaMouvement, PharmaArchive
from models import Vehicule, Entretien, Carburant
from models import CleItem, EmployeCle, AttributionCle, HistoriqueCle
from models import StockProduit, StockMouvement
from models import AnalysePDF
from models import ContratPDF
import json

# ── App factory ───────────────────────────────────────

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_page'


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Non autorise'}), 401
    return redirect(url_for('login_page'))


# ── Creer les tables au demarrage ─────────────────────

with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        import logging
        logging.warning(f"db.create_all() warning: {e}")

    try:
        from sqlalchemy import text, inspect as sa_inspect

        def _get_cols(table):
            try:
                insp = sa_inspect(db.engine)
                return {c['name'] for c in insp.get_columns(table)}
            except Exception:
                return set()

        def _add_col(table, col, col_type):
            try:
                existing = _get_cols(table)
                if col not in existing:
                    db.session.execute(text(f'ALTER TABLE "{table}" ADD COLUMN "{col}" {col_type}'))
                    db.session.commit()
            except Exception:
                db.session.rollback()

        _add_col('users',     'role',          "VARCHAR(20) DEFAULT 'user'")
        _add_col('personnel', 'type_contrat',  "VARCHAR(50) DEFAULT ''")
        _add_col('personnel', 'poste',         "VARCHAR(100) DEFAULT ''")
        _add_col('personnel', 'service',       "VARCHAR(100) DEFAULT ''")
        _add_col('personnel', 'telephone',     "VARCHAR(30) DEFAULT ''")
        _add_col('personnel', 'lieu',          "VARCHAR(100) DEFAULT ''")
        _add_col('personnel', 'date_arrivee',  "VARCHAR(10) DEFAULT ''")
        _add_col('personnel', 'date_depart',   "VARCHAR(10) DEFAULT ''")
        _add_col('unites',    'description',   "VARCHAR(300) DEFAULT ''")
        _add_col('unites',    'emplacement',   "VARCHAR(200) DEFAULT ''")
    except Exception as e:
        import logging
        logging.warning(f"Migration warning: {e}")


@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Erreur serveur interne'}), 500
    return str(e), 500


# ══════════════════════════════════════════════════════
#  PAGES
# ══════════════════════════════════════════════════════

@app.route('/')
@login_required
def index():
    return render_template('app.html')


@app.route('/login')
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('login.html')


# ══════════════════════════════════════════════════════
#  AUTH API
# ══════════════════════════════════════════════════════

@app.route('/auth/register', methods=['POST'])
def auth_register():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '')

    if not name or not email or len(password) < 6:
        return jsonify({'error': 'Tous les champs sont requis (mot de passe min. 6 caracteres).'}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Cet email est deja utilise.'}), 409

    # Premier utilisateur = superadmin
    is_first = User.query.count() == 0
    user = User(name=name, email=email, role='superadmin' if is_first else 'user')
    user.set_password(password)
    db.session.add(user)
    db.session.flush()

    # Creer un etablissement par defaut
    etab = Etablissement(name='Etablissement principal', user_id=user.id, is_current=True)
    db.session.add(etab)
    db.session.commit()

    return jsonify({'ok': True, 'role': user.role}), 201


@app.route('/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '')

    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify({'error': 'Email ou mot de passe incorrect.'}), 401

    login_user(user, remember=True)
    return jsonify({'ok': True, 'name': user.name, 'role': user.role})


@app.route('/auth/logout', methods=['POST'])
@login_required
def auth_logout():
    logout_user()
    return jsonify({'ok': True})


@app.route('/auth/me')
@login_required
def auth_me():
    return jsonify({'id': current_user.id, 'name': current_user.name, 'email': current_user.email, 'role': current_user.role})


# ══════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════

def get_current_etab():
    """Retourne l'etablissement actif de l'utilisateur courant."""
    etab = Etablissement.query.filter_by(user_id=current_user.id, is_current=True).first()
    if not etab:
        etab = Etablissement.query.filter_by(user_id=current_user.id).first()
        if etab:
            etab.is_current = True
            db.session.commit()
    return etab


def user_etabs():
    """Tous les etablissements de l'utilisateur."""
    return Etablissement.query.filter_by(user_id=current_user.id).order_by(Etablissement.id).all()


# ══════════════════════════════════════════════════════
#  API ETABLISSEMENTS
# ══════════════════════════════════════════════════════

@app.route('/api/etablissements', methods=['GET'])
@login_required
def api_get_etablissements():
    etabs = user_etabs()
    current = get_current_etab()
    return jsonify({
        'list': [{'id': e.id, 'name': e.name} for e in etabs],
        'current': {'id': current.id, 'name': current.name} if current else None
    })


@app.route('/api/etablissements', methods=['POST'])
@login_required
def api_add_etablissement():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Nom requis.'}), 400

    existing = Etablissement.query.filter_by(user_id=current_user.id, name=name).first()
    if existing:
        return jsonify({'error': 'Cet etablissement existe deja.'}), 409

    # Desactiver l'ancien courant
    Etablissement.query.filter_by(user_id=current_user.id, is_current=True).update({'is_current': False})
    etab = Etablissement(name=name, user_id=current_user.id, is_current=True)
    db.session.add(etab)
    db.session.commit()
    return jsonify({'ok': True, 'id': etab.id, 'name': etab.name}), 201


@app.route('/api/etablissements/<int:eid>/select', methods=['POST'])
@login_required
def api_select_etablissement(eid):
    etab = Etablissement.query.filter_by(id=eid, user_id=current_user.id).first()
    if not etab:
        return jsonify({'error': 'Introuvable.'}), 404
    Etablissement.query.filter_by(user_id=current_user.id, is_current=True).update({'is_current': False})
    etab.is_current = True
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/etablissements/<int:eid>', methods=['PUT'])
@login_required
def api_rename_etablissement(eid):
    etab = Etablissement.query.filter_by(id=eid, user_id=current_user.id).first()
    if not etab:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Nom requis.'}), 400
    etab.name = name
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/etablissements/<int:eid>', methods=['DELETE'])
@login_required
def api_delete_etablissement(eid):
    etabs = user_etabs()
    if len(etabs) <= 1:
        return jsonify({'error': 'Impossible de supprimer le seul etablissement.'}), 400
    etab = Etablissement.query.filter_by(id=eid, user_id=current_user.id).first()
    if not etab:
        return jsonify({'error': 'Introuvable.'}), 404
    was_current = etab.is_current
    db.session.delete(etab)
    if was_current:
        other = Etablissement.query.filter_by(user_id=current_user.id).first()
        if other:
            other.is_current = True
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  API PRESTATAIRES
# ══════════════════════════════════════════════════════

@app.route('/api/prestataires', methods=['GET'])
@login_required
def api_get_prestataires():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    prestas = Prestataire.query.filter_by(etab_id=etab.id).order_by(Prestataire.nom).all()
    return jsonify([p.to_dict() for p in prestas])


@app.route('/api/prestataires', methods=['POST'])
@login_required
def api_add_prestataire():
    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Aucun etablissement.'}), 400
    data = request.get_json()
    p = Prestataire(
        etab_id=etab.id,
        nom=data.get('nom', ''),
        service=data.get('service', ''),
        contact=data.get('contact', ''),
        email=data.get('email', ''),
        telephone=data.get('telephone', ''),
        date_debut=data.get('date_debut', ''),
        date_fin=data.get('date_fin', ''),
        montant=data.get('montant', 0),
        frequence=data.get('frequence', ''),
        prestations=data.get('prestations', ''),
        notes=data.get('notes', '')
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@app.route('/api/prestataires/<int:pid>', methods=['PUT'])
@login_required
def api_update_prestataire(pid):
    etab = get_current_etab()
    p = Prestataire.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    for field in ['nom', 'service', 'contact', 'email', 'telephone', 'date_debut', 'date_fin', 'montant', 'frequence', 'prestations', 'notes']:
        if field in data:
            setattr(p, field, data[field])
    db.session.commit()
    return jsonify(p.to_dict())


@app.route('/api/prestataires/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_prestataire(pid):
    etab = get_current_etab()
    p = Prestataire.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p:
        return jsonify({'error': 'Introuvable.'}), 404
    # Deplacer en corbeille
    corb = CorbeillePresta(
        etab_id=etab.id, nom=p.nom, service=p.service,
        montant=p.montant, data_json=json.dumps(p.to_dict())
    )
    db.session.add(corb)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})


# Evaluations
@app.route('/api/prestataires/<int:pid>/evaluations', methods=['POST'])
@login_required
def api_add_evaluation(pid):
    etab = get_current_etab()
    p = Prestataire.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    ev = Evaluation(presta_id=pid, date=data.get('date', ''), note=data.get('note', 4), commentaire=data.get('commentaire', ''))
    db.session.add(ev)
    db.session.commit()
    return jsonify(ev.to_dict()), 201


@app.route('/api/evaluations/<int:eid>', methods=['DELETE'])
@login_required
def api_delete_evaluation(eid):
    ev = Evaluation.query.get(eid)
    if not ev:
        return jsonify({'error': 'Introuvable.'}), 404
    db.session.delete(ev)
    db.session.commit()
    return jsonify({'ok': True})


# Corbeille prestas
@app.route('/api/corbeille/prestataires', methods=['GET'])
@login_required
def api_get_corbeille_prestas():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = CorbeillePresta.query.filter_by(etab_id=etab.id).order_by(CorbeillePresta.deleted_at.desc()).all()
    return jsonify([c.to_dict() for c in items])


@app.route('/api/corbeille/prestataires/<int:cid>/restore', methods=['POST'])
@login_required
def api_restore_prestataire(cid):
    etab = get_current_etab()
    c = CorbeillePresta.query.filter_by(id=cid, etab_id=etab.id).first()
    if not c:
        return jsonify({'error': 'Introuvable.'}), 404
    data = json.loads(c.data_json)
    p = Prestataire(
        etab_id=etab.id, nom=data.get('nom', ''), service=data.get('service', ''),
        contact=data.get('contact', ''), email=data.get('email', ''),
        telephone=data.get('telephone', ''), date_debut=data.get('date_debut', ''),
        date_fin=data.get('date_fin', ''), montant=data.get('montant', 0),
        frequence=data.get('frequence', ''), prestations=data.get('prestations', ''),
        notes=data.get('notes', '')
    )
    db.session.add(p)
    db.session.delete(c)
    db.session.commit()
    return jsonify(p.to_dict()), 201


# ══════════════════════════════════════════════════════
#  API ACTIFS — Personnel
# ══════════════════════════════════════════════════════

@app.route('/api/personnel', methods=['GET'])
@login_required
def api_get_personnel():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = Personnel.query.filter_by(etab_id=etab.id).order_by(Personnel.nom).all()
    return jsonify([p.to_dict() for p in items])


@app.route('/api/personnel', methods=['POST'])
@login_required
def api_add_personnel():
    etab = get_current_etab()
    data = request.get_json()
    p = Personnel(etab_id=etab.id, nom=data.get('nom', ''), prenom=data.get('prenom', ''),
                  type_contrat=data.get('type_contrat', ''), poste=data.get('poste', ''),
                  service=data.get('service', ''), telephone=data.get('telephone', ''),
                  lieu=data.get('lieu', ''), date_arrivee=data.get('date_arrivee', ''),
                  date_depart=data.get('date_depart', ''))
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@app.route('/api/personnel/<int:pid>', methods=['PUT'])
@login_required
def api_update_personnel(pid):
    etab = get_current_etab()
    p = Personnel.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    for f in ['nom', 'prenom', 'type_contrat', 'poste', 'service', 'telephone', 'lieu', 'date_arrivee', 'date_depart']:
        if f in data:
            setattr(p, f, data[f])
    db.session.commit()
    return jsonify(p.to_dict())


@app.route('/api/personnel/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_personnel(pid):
    etab = get_current_etab()
    p = Personnel.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p:
        return jsonify({'error': 'Introuvable.'}), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/personnel/export-excel', methods=['GET'])
@login_required
def api_export_personnel_excel():
    """Export Personnel → fichier Excel téléchargeable."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl non installé.'}), 500
    etab = get_current_etab()
    items = Personnel.query.filter_by(etab_id=etab.id).order_by(Personnel.nom).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Personnel'
    headers = ['Nom', 'Prénom', 'Poste', 'Service', 'Type contrat', 'Téléphone', 'Date entrée', 'Date sortie']
    col_widths = [20, 20, 22, 20, 18, 16, 14, 14]
    hdr_fill = PatternFill('solid', fgColor='1E3A8A')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = w
    CONTRACT_OPTIONS = ['', 'CDI', 'CDD', 'Intérim', 'Stage', 'Apprentissage', 'Bénévolat', 'Intervenant ext', 'Autre']
    for row_i, p in enumerate(items, 2):
        ws.cell(row=row_i, column=1, value=p.nom)
        ws.cell(row=row_i, column=2, value=p.prenom)
        ws.cell(row=row_i, column=3, value=p.poste)
        ws.cell(row=row_i, column=4, value=p.service)
        ws.cell(row=row_i, column=5, value=p.type_contrat)
        ws.cell(row=row_i, column=6, value=p.telephone)
        ws.cell(row=row_i, column=7, value=p.date_arrivee)
        ws.cell(row=row_i, column=8, value=p.date_depart)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='personnel.xlsx')


@app.route('/api/config/template', methods=['GET'])
@login_required
def api_config_template():
    """Génère un fichier Excel modèle avec 3 lieux et 3 personnel."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl non installé.'}), 500

    wb = openpyxl.Workbook()

    # ── Feuille Personnel ──
    ws_pers = wb.active
    ws_pers.title = 'Personnel'
    pers_headers = ['Nom', 'Prénom', 'Poste', 'Service', 'Type contrat', 'Téléphone', 'Date entrée', 'Date sortie']
    pers_widths  = [18, 18, 22, 20, 18, 16, 14, 14]
    fill_p = PatternFill('solid', fgColor='1E3A8A')
    font_h = Font(bold=True, color='FFFFFF', size=11)
    for ci, (h, w) in enumerate(zip(pers_headers, pers_widths), 1):
        c = ws_pers.cell(row=1, column=ci, value=h)
        c.fill = fill_p; c.font = font_h
        c.alignment = Alignment(horizontal='center')
        ws_pers.column_dimensions[c.column_letter].width = w
    pers_rows = [
        ('DUPONT',  'Marie',  'Éducatrice',   'Unité A', 'CDI',  '06 11 22 33 44', '2020-09-01', ''),
        ('MARTIN',  'Paul',   'Infirmier',    'Soins',   'CDI',  '06 55 66 77 88', '2019-03-15', ''),
        ('PETIT',   'Lucas',  'Agent d\'entretien', 'Technique', 'CDD', '06 99 00 11 22', '2024-01-10', '2025-01-09'),
    ]
    for ri, row in enumerate(pers_rows, 2):
        for ci, val in enumerate(row, 1):
            ws_pers.cell(row=ri, column=ci, value=val)

    # ── Feuille Lieux ──
    ws_lieux = wb.create_sheet('Lieux')
    lieux_headers = ['Nom', 'Description', 'Emplacement']
    lieux_widths  = [28, 35, 25]
    fill_l = PatternFill('solid', fgColor='166534')
    for ci, (h, w) in enumerate(zip(lieux_headers, lieux_widths), 1):
        c = ws_lieux.cell(row=1, column=ci, value=h)
        c.fill = fill_l; c.font = font_h
        c.alignment = Alignment(horizontal='center')
        ws_lieux.column_dimensions[c.column_letter].width = w
    lieux_rows = [
        ('Les Choucas',  'Unité de vie principale',   'Bâtiment A – RDC'),
        ('Infirmerie',   'Salle de soins',             'Bâtiment B – 1er étage'),
        ('Cuisine',      'Cuisine collective',         'Bâtiment A – Sous-sol'),
    ]
    for ri, row in enumerate(lieux_rows, 2):
        for ci, val in enumerate(row, 1):
            ws_lieux.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='modele_config.xlsx')


@app.route('/api/personnel/import-excel', methods=['POST'])
@login_required
def api_import_personnel_excel():
    """Import Personnel depuis fichier Excel (ajoute, ne remplace pas)."""
    import io
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl non installé.'}), 500
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Aucun fichier.'}), 400
    etab = get_current_etab()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            nom = str(row[0]).strip()
            if not nom:
                continue
            r = tuple(row) + (None,) * 8  # pad pour eviter index out of range
            prenom      = str(r[1] or '').strip()
            poste       = str(r[2] or '').strip()
            service     = str(r[3] or '').strip()
            type_contrat= str(r[4] or '').strip()
            telephone   = str(r[5] or '').strip()
            date_arr    = str(r[6] or '').strip()
            date_dep    = str(r[7] or '').strip()
            existing = Personnel.query.filter_by(etab_id=etab.id, nom=nom, prenom=prenom).first()
            if existing:
                existing.poste = poste or existing.poste
                existing.service = service or existing.service
                existing.type_contrat = type_contrat or existing.type_contrat
                existing.telephone = telephone or existing.telephone
                existing.date_arrivee = date_arr or existing.date_arrivee
                existing.date_depart = date_dep or existing.date_depart
            else:
                db.session.add(Personnel(etab_id=etab.id, nom=nom, prenom=prenom,
                    poste=poste, service=service, type_contrat=type_contrat,
                    telephone=telephone, date_arrivee=date_arr, date_depart=date_dep))
                added += 1
        db.session.commit()
        return jsonify({'ok': True, 'added': added})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ══════════════════════════════════════════════════════
#  API ACTIFS — Unites
# ══════════════════════════════════════════════════════

@app.route('/api/unites', methods=['GET'])
@login_required
def api_get_unites():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = Unite.query.filter_by(etab_id=etab.id).order_by(Unite.nom).all()
    return jsonify([u.to_dict() for u in items])


@app.route('/api/unites', methods=['POST'])
@login_required
def api_add_unite():
    etab = get_current_etab()
    data = request.get_json()
    u = Unite(etab_id=etab.id, nom=data.get('nom', ''), description=data.get('description', ''),
              emplacement=data.get('emplacement', ''))
    db.session.add(u)
    db.session.commit()
    return jsonify(u.to_dict()), 201


@app.route('/api/unites/<int:uid>', methods=['PUT'])
@login_required
def api_update_unite(uid):
    etab = get_current_etab()
    u = Unite.query.filter_by(id=uid, etab_id=etab.id).first()
    if not u:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    for f in ['nom', 'description', 'emplacement']:
        if f in data:
            setattr(u, f, data[f])
    db.session.commit()
    return jsonify(u.to_dict())


@app.route('/api/unites/<int:uid>', methods=['DELETE'])
@login_required
def api_delete_unite(uid):
    etab = get_current_etab()
    u = Unite.query.filter_by(id=uid, etab_id=etab.id).first()
    if not u:
        return jsonify({'error': 'Introuvable.'}), 404
    db.session.delete(u)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  API ACTIFS — Materiels
# ══════════════════════════════════════════════════════

@app.route('/api/materiels', methods=['GET'])
@login_required
def api_get_materiels():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = Materiel.query.filter_by(etab_id=etab.id).order_by(Materiel.nom).all()
    return jsonify([m.to_dict() for m in items])


@app.route('/api/materiels/scan-facture', methods=['POST'])
@login_required
def scan_facture_ia():
    import os as _os, tempfile as _tmp, json as _json, re as _re, urllib.request, base64 as _b64

    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    fname = f.filename.lower()
    if not (fname.endswith('.pdf') or fname.endswith('.jpg') or fname.endswith('.jpeg') or fname.endswith('.png')):
        return jsonify({'error': 'Format accepté : PDF, JPG, PNG'}), 400

    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API Gemini requise'}), 400

    suffix = '.' + fname.rsplit('.', 1)[-1]
    with _tmp.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        prompt = (
            "Analyse cette facture d'achat et extrait les informations suivantes en JSON strict, "
            "sans markdown ni texte autour :\n"
            '{"nom": "désignation ou nom du matériel/équipement acheté (3-8 mots max)", '
            '"reference": "numéro de facture ou bon de commande si présent, sinon vide", '
            '"numero_serie": "numéro de série (labels: S/N, Série, Serial, N° série, SN) — vide si absent", '
            '"type_materiel": "type parmi: Audiovisuel, Climatisation, Electroménager, Équipement, Informatique, Médical, Mobilier, Outillage, Sécurité, Téléphonie, Travaux, Véhicule, Autre", '
            '"date_achat": "YYYY-MM-DD", '
            '"cout": 0, '
            '"notes": "résumé en 1-2 phrases"}\n'
            "Règles : date_achat toujours en YYYY-MM-DD. cout = montant total TTC (si HT seulement, multiplier par 1.20). "
            "Si valeur absente : chaîne vide ou 0."
        )

        if suffix == '.pdf':
            texte = _extraire_texte_pdf(tmp_path)
            if texte:
                body = _json.dumps({
                    "contents": [{"parts": [{"text": prompt + "\n\nTexte de la facture :\n" + texte[:40000]}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
            else:
                # PDF scanné → Vision Gemini
                pdf_b64 = _b64.b64encode(open(tmp_path, 'rb').read()).decode('utf-8')
                body = _json.dumps({
                    "contents": [{"parts": [
                        {"text": prompt + "\n\n[Analyse la facture visible dans ce PDF scanné]"},
                        {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                    ]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
        else:
            with open(tmp_path, 'rb') as img_f:
                img_data = _b64.b64encode(img_f.read()).decode('utf-8')
            mime = 'image/jpeg' if suffix in ('.jpg', '.jpeg') else 'image/png'
            body = _json.dumps({
                "contents": [{"parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime, "data": img_data}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=60) as r:
            result = _json.loads(r.read().decode("utf-8"))
        texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
        texte_rep = _re.sub(r"```json\s*", "", texte_rep)
        texte_rep = _re.sub(r"```\s*", "", texte_rep)
        data = _parse_gemini_json(texte_rep)
        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': f'Erreur scan : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.route('/api/materiels', methods=['POST'])
@login_required
def api_add_materiel():
    etab = get_current_etab()
    data = request.get_json()
    m = Materiel(etab_id=etab.id, nom=data.get('nom', ''), reference=data.get('reference', ''),
                 type_materiel=data.get('type_materiel', ''), date_achat=data.get('date_achat', ''),
                 cout=data.get('cout', 0), duree_amortissement=data.get('duree_amortissement', 5),
                 statut=data.get('statut', 'En service'), attribue_a=data.get('attribue_a', ''),
                 notes=data.get('notes', ''))
    db.session.add(m)
    db.session.commit()
    return jsonify(m.to_dict()), 201


@app.route('/api/materiels/<int:mid>', methods=['PUT'])
@login_required
def api_update_materiel(mid):
    etab = get_current_etab()
    m = Materiel.query.filter_by(id=mid, etab_id=etab.id).first()
    if not m:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    for f in ['nom', 'reference', 'type_materiel', 'date_achat', 'cout', 'duree_amortissement', 'statut', 'attribue_a', 'notes']:
        if f in data:
            setattr(m, f, data[f])
    db.session.commit()
    return jsonify(m.to_dict())


@app.route('/api/materiels/<int:mid>', methods=['DELETE'])
@login_required
def api_delete_materiel(mid):
    etab = get_current_etab()
    m = Materiel.query.filter_by(id=mid, etab_id=etab.id).first()
    if not m:
        return jsonify({'error': 'Introuvable.'}), 404
    db.session.delete(m)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  API PHARMACIE — Stock
# ══════════════════════════════════════════════════════

@app.route('/api/pharmacie/stock', methods=['GET'])
@login_required
def api_get_pharma_stock():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = PharmaStock.query.filter_by(etab_id=etab.id).order_by(PharmaStock.nom_medicament).all()
    return jsonify([s.to_dict() for s in items])


@app.route('/api/pharmacie/stock', methods=['POST'])
@login_required
def api_add_pharma_stock():
    etab = get_current_etab()
    data = request.get_json()
    nom = data.get('nom_medicament', '')
    lot = data.get('lot', '')
    per = data.get('date_peremption', '')

    # Verifier duplicat
    existing = PharmaStock.query.filter_by(etab_id=etab.id, nom_medicament=nom, lot=lot, date_peremption=per).first()
    if existing:
        existing.quantite += data.get('quantite', 1)
        db.session.commit()
        s = existing
    else:
        s = PharmaStock(
            etab_id=etab.id, nom_medicament=nom, lot=lot, date_peremption=per,
            quantite=data.get('quantite', 1), stock_minimum=data.get('stock_minimum', 0),
            emplacement=data.get('emplacement', ''), personne_entree=data.get('personne_entree', ''),
            date_ajout=data.get('date_ajout', ''), derniere_sortie=''
        )
        db.session.add(s)
        db.session.commit()

    # Mouvement reception
    mv = PharmaMouvement(
        etab_id=etab.id, type='reception', nom_medicament=nom,
        quantite=data.get('quantite', 1), personne=data.get('personne_entree', ''),
        role=data.get('personne_entree', ''), date_mouvement=data.get('date_ajout', ''),
        heure=data.get('heure', '')
    )
    db.session.add(mv)
    db.session.commit()
    return jsonify(s.to_dict()), 201


@app.route('/api/pharmacie/stock/<int:sid>', methods=['PUT'])
@login_required
def api_update_pharma_stock(sid):
    etab = get_current_etab()
    s = PharmaStock.query.filter_by(id=sid, etab_id=etab.id).first()
    if not s:
        return jsonify({'error': 'Introuvable.'}), 404
    data = request.get_json()
    for f in ['nom_medicament', 'lot', 'date_peremption', 'quantite', 'stock_minimum', 'emplacement', 'personne_entree', 'date_ajout']:
        if f in data:
            setattr(s, f, data[f])
    db.session.commit()
    return jsonify(s.to_dict())


@app.route('/api/pharmacie/stock/<int:sid>', methods=['DELETE'])
@login_required
def api_delete_pharma_stock(sid):
    etab = get_current_etab()
    s = PharmaStock.query.filter_by(id=sid, etab_id=etab.id).first()
    if not s:
        return jsonify({'error': 'Introuvable.'}), 404
    # Archive
    a = PharmaArchive(etab_id=etab.id, nom_medicament=s.nom_medicament, lot=s.lot,
                      date_peremption=s.date_peremption, quantite=s.quantite,
                      emplacement=s.emplacement, date_suppression=data_today())
    db.session.add(a)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'ok': True})


# Sortie
@app.route('/api/pharmacie/sortie', methods=['POST'])
@login_required
def api_pharma_sortie():
    etab = get_current_etab()
    data = request.get_json()
    nom = data.get('nom_medicament', '')
    qte = data.get('quantite', 0)
    personne = data.get('personne', '')

    # FIFO
    candidates = PharmaStock.query.filter_by(etab_id=etab.id, nom_medicament=nom)\
        .filter(PharmaStock.quantite > 0).order_by(PharmaStock.date_peremption).all()

    total_dispo = sum(c.quantite for c in candidates)
    if qte > total_dispo:
        return jsonify({'error': f'Stock insuffisant (disponible : {total_dispo}).'}), 400

    restant = qte
    for c in candidates:
        if restant <= 0:
            break
        retire = min(restant, c.quantite)
        c.quantite -= retire
        c.derniere_sortie = data_today()
        restant -= retire

    # Supprimer stock a 0
    PharmaStock.query.filter_by(etab_id=etab.id, nom_medicament=nom).filter(PharmaStock.quantite <= 0).delete()

    # Mouvement
    mv = PharmaMouvement(
        etab_id=etab.id, type='sortie', nom_medicament=nom,
        quantite=qte, personne=personne, role=personne,
        date_mouvement=data_today(), heure=data.get('heure', '')
    )
    db.session.add(mv)
    db.session.commit()
    return jsonify({'ok': True})


# Mouvements
@app.route('/api/pharmacie/mouvements', methods=['GET'])
@login_required
def api_get_pharma_mouvements():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = PharmaMouvement.query.filter_by(etab_id=etab.id)\
        .order_by(PharmaMouvement.date_mouvement.desc(), PharmaMouvement.id.desc()).all()
    return jsonify([m.to_dict() for m in items])


# Archive
@app.route('/api/pharmacie/archive', methods=['GET'])
@login_required
def api_get_pharma_archive():
    etab = get_current_etab()
    if not etab:
        return jsonify([])
    items = PharmaArchive.query.filter_by(etab_id=etab.id).all()
    return jsonify([a.to_dict() for a in items])


@app.route('/api/pharmacie/archive/<int:aid>/restore', methods=['POST'])
@login_required
def api_restore_pharma(aid):
    etab = get_current_etab()
    a = PharmaArchive.query.filter_by(id=aid, etab_id=etab.id).first()
    if not a:
        return jsonify({'error': 'Introuvable.'}), 404
    s = PharmaStock(etab_id=etab.id, nom_medicament=a.nom_medicament, lot=a.lot,
                    date_peremption=a.date_peremption, quantite=a.quantite,
                    emplacement=a.emplacement, date_ajout=data_today())
    db.session.add(s)
    db.session.delete(a)
    db.session.commit()
    return jsonify(s.to_dict()), 201


# Retirer perimes
@app.route('/api/pharmacie/retirer-perimes', methods=['POST'])
@login_required
def api_pharma_retirer_perimes():
    from datetime import date
    etab = get_current_etab()
    today_str = date.today().isoformat()
    perimes = PharmaStock.query.filter_by(etab_id=etab.id).filter(PharmaStock.date_peremption < today_str).all()
    count = len(perimes)
    for p in perimes:
        a = PharmaArchive(etab_id=etab.id, nom_medicament=p.nom_medicament, lot=p.lot,
                          date_peremption=p.date_peremption, quantite=p.quantite,
                          emplacement=p.emplacement, date_suppression=today_str)
        db.session.add(a)
        db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True, 'count': count})


# ══════════════════════════════════════════════════════
#  API PHARMACIE — BDPM (base officielle medicaments)
# ══════════════════════════════════════════════════════

BDPM_URL_CIS = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_bdpm.txt"
BDPM_URL_CIP = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_CIP_bdpm.txt"

# Table BDPM en memoire (partagee entre requetes)
_bdpm_cache = {}  # cip13 -> nom_medicament
_bdpm_loaded = False


def _bdpm_load():
    """Charge la BDPM en memoire depuis les fichiers gouvernementaux."""
    global _bdpm_cache, _bdpm_loaded
    import urllib.request
    import ssl

    try:
        ctx = ssl.create_default_context()
    except Exception:
        ctx = ssl._create_unverified_context()

    # 1. CIS_bdpm.txt : code CIS -> nom medicament
    cis_noms = {}
    try:
        req = urllib.request.Request(BDPM_URL_CIS, headers={'User-Agent': 'BusinessManager/1.0'})
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            text = resp.read().decode('latin-1')
        for line in text.strip().split('\n'):
            parts = line.split('\t')
            if len(parts) >= 2:
                cis = parts[0].strip()
                nom = parts[1].strip()
                cis_noms[cis] = nom
    except Exception as e:
        return False, f"Erreur telechargement CIS: {e}"

    # 2. CIS_CIP_bdpm.txt : correspondance CIP13 -> CIS
    # Format : 0:cis  1:cip7  2:nom_presentation  3:statut_adm
    #          4:etat_commerc  5:date_decl  6:cip13  7:agrement
    count = 0
    try:
        req = urllib.request.Request(BDPM_URL_CIP, headers={'User-Agent': 'BusinessManager/1.0'})
        with urllib.request.urlopen(req, context=ctx, timeout=60) as resp:
            text = resp.read().decode('latin-1')
        for line in text.strip().split('\n'):
            parts = line.split('\t')
            if len(parts) >= 7:
                cis = parts[0].strip()
                cip13 = parts[6].strip()
                nom_pres = parts[2].strip() if len(parts) > 2 else ''
                nom_cis = cis_noms.get(cis, '')
                nom = nom_cis or nom_pres
                if cip13 and nom:
                    _bdpm_cache[cip13] = nom
                    count += 1
    except Exception as e:
        return False, f"Erreur telechargement CIP: {e}"

    _bdpm_loaded = True
    return True, f"{count} medicaments indexes"


@app.route('/api/pharmacie/bdpm/status', methods=['GET'])
@login_required
def api_bdpm_status():
    return jsonify({'loaded': _bdpm_loaded, 'count': len(_bdpm_cache)})


@app.route('/api/pharmacie/bdpm/download', methods=['POST'])
@login_required
def api_bdpm_download():
    ok, msg = _bdpm_load()
    if ok:
        return jsonify({'ok': True, 'message': msg, 'count': len(_bdpm_cache)})
    return jsonify({'error': msg}), 500


@app.route('/api/pharmacie/bdpm/lookup/<cip13>', methods=['GET'])
@login_required
def api_bdpm_lookup(cip13):
    cip13 = cip13.strip()
    nom = _bdpm_cache.get(cip13)
    if nom:
        return jsonify({'found': True, 'nom': nom, 'cip13': cip13})
    return jsonify({'found': False, 'cip13': cip13})


@app.route('/api/pharmacie/bdpm/search', methods=['GET'])
@login_required
def api_bdpm_search():
    q = request.args.get('q', '').strip().lower()
    if len(q) < 2:
        return jsonify([])
    results = []
    for cip, nom in _bdpm_cache.items():
        if q in nom.lower():
            results.append({'cip13': cip, 'nom': nom})
            if len(results) >= 20:
                break
    return jsonify(results)


# ══════════════════════════════════════════════════════
#  API ADMIN
# ══════════════════════════════════════════════════════

def require_admin(f):
    """Decorateur : necessite le role admin ou superadmin."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_admin():
            return jsonify({'error': 'Acces refuse. Droits administrateur requis.'}), 403
        return f(*args, **kwargs)
    return decorated


def require_superadmin(f):
    """Decorateur : necessite le role superadmin."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_superadmin():
            return jsonify({'error': 'Acces refuse. Droits super-administrateur requis.'}), 403
        return f(*args, **kwargs)
    return decorated


@app.route('/api/admin/users', methods=['GET'])
@login_required
@require_admin
def api_admin_list_users():
    """Liste tous les utilisateurs (admin+ seulement)."""
    users = User.query.order_by(User.created_at).all()
    result = []
    for u in users:
        etabs = Etablissement.query.filter_by(user_id=u.id).all()
        result.append({
            **u.to_dict(),
            'etablissements': [{'id': e.id, 'name': e.name} for e in etabs]
        })
    return jsonify(result)


@app.route('/api/admin/users', methods=['POST'])
@login_required
@require_admin
def api_admin_create_user():
    """Creer un utilisateur (admin+ seulement)."""
    data = request.get_json()
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', 'user')

    if not name or not email or len(password) < 6:
        return jsonify({'error': 'Nom, email et mot de passe (min 6 car.) requis.'}), 400

    # Seul superadmin peut creer des admin/superadmin
    if role in ('admin', 'superadmin') and not current_user.is_superadmin():
        return jsonify({'error': 'Seul un super-admin peut creer des administrateurs.'}), 403

    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Cet email est deja utilise.'}), 409

    user = User(name=name, email=email, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.flush()

    # Creer un etablissement par defaut
    etab = Etablissement(name='Etablissement principal', user_id=user.id, is_current=True)
    db.session.add(etab)
    db.session.commit()

    return jsonify(user.to_dict()), 201


@app.route('/api/admin/users/<int:uid>', methods=['PUT'])
@login_required
@require_admin
def api_admin_update_user(uid):
    """Modifier un utilisateur (role, nom, etc.)."""
    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Utilisateur introuvable.'}), 404

    data = request.get_json()

    # Changer le role : superadmin only
    new_role = data.get('role')
    if new_role and new_role != user.role:
        if not current_user.is_superadmin():
            return jsonify({'error': 'Seul un super-admin peut changer les roles.'}), 403
        if new_role not in ('user', 'admin', 'superadmin'):
            return jsonify({'error': 'Role invalide.'}), 400
        user.role = new_role

    if 'name' in data:
        user.name = data['name'].strip()
    if 'email' in data:
        existing = User.query.filter_by(email=data['email'].strip().lower()).first()
        if existing and existing.id != uid:
            return jsonify({'error': 'Email deja utilise.'}), 409
        user.email = data['email'].strip().lower()
    if data.get('password') and len(data['password']) >= 6:
        user.set_password(data['password'])

    db.session.commit()
    return jsonify(user.to_dict())


@app.route('/api/admin/users/<int:uid>', methods=['DELETE'])
@login_required
@require_superadmin
def api_admin_delete_user(uid):
    """Supprimer un utilisateur (superadmin only)."""
    if uid == current_user.id:
        return jsonify({'error': 'Impossible de se supprimer soi-meme.'}), 400
    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Utilisateur introuvable.'}), 404
    # Supprimer ses etablissements (cascade supprime tout le reste)
    for etab in user.etablissements:
        db.session.delete(etab)
    db.session.delete(user)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/admin/users/<int:uid>/etabs', methods=['POST'])
@login_required
@require_admin
def api_admin_assign_etab(uid):
    """Creer un etablissement pour un utilisateur."""
    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Utilisateur introuvable.'}), 404
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Nom requis.'}), 400
    etab = Etablissement(name=name, user_id=uid, is_current=False)
    db.session.add(etab)
    db.session.commit()
    return jsonify({'ok': True, 'id': etab.id, 'name': etab.name}), 201


# ══════════════════════════════════════════════════════
#  API SEED DEMO
# ══════════════════════════════════════════════════════

@app.route('/api/seed-demo', methods=['POST'])
@login_required
def api_seed_demo():
    """Peuple l'etablissement actif avec des donnees de demonstration."""
    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Aucun etablissement.'}), 400

    today = data_today()

    # -- Prestataires demo --
    prestas_demo = [
        {'nom': 'CleanPro Services', 'service': 'Nettoyage', 'contact': 'Marie Dupont',
         'email': 'contact@cleanpro.fr', 'telephone': '01 23 45 67 89',
         'date_debut': '2024-01-01', 'date_fin': '2026-12-31',
         'montant': 2500, 'frequence': 'Mensuel',
         'prestations': 'Nettoyage quotidien des locaux, bureaux et sanitaires',
         'notes': 'Intervention le matin entre 7h et 9h'},
        {'nom': 'TechMaint SARL', 'service': 'Maintenance', 'contact': 'Jean Martin',
         'email': 'jean.martin@techmaint.fr', 'telephone': '06 12 34 56 78',
         'date_debut': '2023-06-01', 'date_fin': '2025-05-31',
         'montant': 4800, 'frequence': 'Trimestriel',
         'prestations': 'Maintenance préventive équipements électriques et climatisation',
         'notes': 'Astreinte 24h/24 incluse'},
        {'nom': 'RestauCollect', 'service': 'Restauration', 'contact': 'Sophie Bernard',
         'email': 'sophie@restaucollect.com', 'telephone': '01 98 76 54 32',
         'date_debut': '2024-09-01', 'date_fin': '2025-08-31',
         'montant': 12000, 'frequence': 'Mensuel',
         'prestations': 'Fourniture et distribution des repas midi et soir',
         'notes': 'Menu adapté régime sans porc disponible'},
        {'nom': 'VerdiSpace', 'service': 'Espaces verts', 'contact': 'Paul Leblanc',
         'email': 'contact@verdispace.fr', 'telephone': '06 87 65 43 21',
         'date_debut': '2024-03-01', 'date_fin': '2025-02-28',
         'montant': 800, 'frequence': 'Mensuel',
         'prestations': 'Entretien jardins, tonte pelouses, taille haies',
         'notes': 'Passage tous les 15 jours au printemps'},
        {'nom': 'SecurGuard', 'service': 'Sécurité', 'contact': 'Ahmed Ould',
         'email': 'a.ould@securguard.fr', 'telephone': '01 45 67 89 01',
         'date_debut': '2023-01-01', 'date_fin': '2025-12-31',
         'montant': 6500, 'frequence': 'Mensuel',
         'prestations': 'Gardiennage nuit et week-end, rondes de sécurité',
         'notes': 'Agent armé selon protocole établissement'},
    ]
    for d in prestas_demo:
        if not Prestataire.query.filter_by(etab_id=etab.id, nom=d['nom']).first():
            db.session.add(Prestataire(etab_id=etab.id, **d))

    # -- Personnel demo --
    pers_demo = [
        {'nom': 'Durand', 'prenom': 'Claire', 'type_contrat': 'CDI', 'poste': 'Directrice',
         'lieu': 'Bâtiment A', 'date_arrivee': '2018-09-01', 'date_depart': ''},
        {'nom': 'Lambert', 'prenom': 'Pierre', 'type_contrat': 'CDI', 'poste': 'Infirmier',
         'lieu': 'Infirmerie', 'date_arrivee': '2020-02-15', 'date_depart': ''},
        {'nom': 'Moreau', 'prenom': 'Julie', 'type_contrat': 'CDD', 'poste': 'Educatrice',
         'lieu': 'Bâtiment B', 'date_arrivee': '2024-01-10', 'date_depart': '2025-06-30'},
        {'nom': 'Garcia', 'prenom': 'Carlos', 'type_contrat': 'CDI', 'poste': 'Agent d\'entretien',
         'lieu': 'Bâtiment C', 'date_arrivee': '2019-05-01', 'date_depart': ''},
    ]
    for d in pers_demo:
        if not Personnel.query.filter_by(etab_id=etab.id, nom=d['nom'], prenom=d['prenom']).first():
            db.session.add(Personnel(etab_id=etab.id, **d))

    # -- Unites demo --
    unites_demo = [
        {'nom': 'Groupe Tournesol', 'description': 'Groupe des 3-5 ans', 'emplacement': 'Bâtiment A - RDC'},
        {'nom': 'Groupe Papillons', 'description': 'Groupe des 6-10 ans', 'emplacement': 'Bâtiment A - 1er'},
        {'nom': 'Groupe Dauphins', 'description': 'Groupe des 11-15 ans', 'emplacement': 'Bâtiment B'},
        {'nom': 'Unité Autonomie', 'description': 'Résidence adultes autonomes', 'emplacement': 'Bâtiment C'},
    ]
    for d in unites_demo:
        if not Unite.query.filter_by(etab_id=etab.id, nom=d['nom']).first():
            db.session.add(Unite(etab_id=etab.id, **d))

    db.session.flush()

    # -- Materiels demo --
    mats_demo = [
        {'nom': 'Ordinateur portable HP', 'reference': 'HP-2024-001', 'type_materiel': 'Informatique',
         'date_achat': '2024-01-15', 'cout': 899, 'duree_amortissement': 5,
         'statut': 'En service', 'attribue_a': 'Claire Durand', 'notes': 'Directrice'},
        {'nom': 'Climatiseur Daikin Salle A', 'reference': 'DAIKIN-001', 'type_materiel': 'Équipement',
         'date_achat': '2022-06-01', 'cout': 1800, 'duree_amortissement': 10,
         'statut': 'En service', 'attribue_a': '', 'notes': 'Révision annuelle prévue juin'},
        {'nom': 'Véhicule Renault Kangoo', 'reference': 'AA-123-BB', 'type_materiel': 'Véhicule',
         'date_achat': '2021-03-10', 'cout': 18500, 'duree_amortissement': 7,
         'statut': 'En service', 'attribue_a': 'Carlos Garcia', 'notes': 'Contrôle technique mars 2025'},
        {'nom': 'Table de réunion 10 places', 'reference': 'MOB-007', 'type_materiel': 'Mobilier',
         'date_achat': '2020-09-01', 'cout': 650, 'duree_amortissement': 10,
         'statut': 'En service', 'attribue_a': '', 'notes': 'Salle de réunion RDC'},
        {'nom': 'Lave-linge professionnel', 'reference': 'MIELE-PRO-01', 'type_materiel': 'Équipement',
         'date_achat': '2019-11-20', 'cout': 3200, 'duree_amortissement': 8,
         'statut': 'En maintenance', 'attribue_a': '', 'notes': 'Panne courroie — réparation en cours'},
    ]
    for d in mats_demo:
        if not Materiel.query.filter_by(etab_id=etab.id, nom=d['nom']).first():
            db.session.add(Materiel(etab_id=etab.id, **d))

    # -- Pharmacie stock demo --
    from datetime import date
    pharma_demo = [
        {'nom_medicament': 'Doliprane 1000mg', 'lot': 'LOT-2024-001', 'date_peremption': '2026-06-30',
         'quantite': 48, 'stock_minimum': 10, 'emplacement': 'Armoire A - Rayon 1',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2024-10-01'},
        {'nom_medicament': 'Doliprane 1000mg', 'lot': 'LOT-2025-012', 'date_peremption': '2027-03-31',
         'quantite': 24, 'stock_minimum': 10, 'emplacement': 'Armoire A - Rayon 1',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2025-01-15'},
        {'nom_medicament': 'Ibuprofene 400mg', 'lot': 'IBU-2024-789', 'date_peremption': '2026-09-30',
         'quantite': 30, 'stock_minimum': 5, 'emplacement': 'Armoire A - Rayon 2',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2024-11-01'},
        {'nom_medicament': 'Amoxicilline 500mg', 'lot': 'AMX-2025-003', 'date_peremption': '2026-12-31',
         'quantite': 6, 'stock_minimum': 5, 'emplacement': 'Armoire B - Rayon 1',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2025-02-10'},
        {'nom_medicament': 'Ventoline 100mcg', 'lot': 'VENT-2024-441', 'date_peremption': '2025-08-31',
         'quantite': 3, 'stock_minimum': 2, 'emplacement': 'Armoire Urgences',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2024-09-01'},
        {'nom_medicament': 'Dafalgan 500mg', 'lot': 'DAF-2025-110', 'date_peremption': '2027-01-31',
         'quantite': 2, 'stock_minimum': 5, 'emplacement': 'Armoire A - Rayon 1',
         'personne_entree': 'Pierre Lambert', 'date_ajout': '2025-03-01'},
    ]
    for d in pharma_demo:
        existing = PharmaStock.query.filter_by(etab_id=etab.id, nom_medicament=d['nom_medicament'], lot=d['lot']).first()
        if not existing:
            db.session.add(PharmaStock(etab_id=etab.id, **d))
            mv = PharmaMouvement(etab_id=etab.id, type='reception', nom_medicament=d['nom_medicament'],
                                 quantite=d['quantite'], personne=d['personne_entree'],
                                 role='Infirmier', date_mouvement=d['date_ajout'], heure='09:00')
            db.session.add(mv)

    # Evaluations demo pour CleanPro
    db.session.flush()
    cleanpro = Prestataire.query.filter_by(etab_id=etab.id, nom='CleanPro Services').first()
    if cleanpro and not Evaluation.query.filter_by(presta_id=cleanpro.id).first():
        for ev in [
            {'date': '2024-03-15', 'note': 4, 'commentaire': 'Bon travail, quelques oublis sur les sanitaires du 2ème'},
            {'date': '2024-06-20', 'note': 5, 'commentaire': 'Excellent ce trimestre, aucun retour négatif'},
            {'date': '2024-09-10', 'note': 3, 'commentaire': 'Remplacement agent peu efficace en août'},
        ]:
            db.session.add(Evaluation(presta_id=cleanpro.id, **ev))

    db.session.commit()
    return jsonify({'ok': True, 'message': 'Donnees de demonstration chargees avec succes.'})


# ══════════════════════════════════════════════════════
#  ANALYSE PDF
# ══════════════════════════════════════════════════════

PROMPT_INSPECTION = """
Tu es un expert en sécurité bâtiment et réglementation ERP. Analyse ce rapport d'inspection.

CLASSIFICATIONS : Alerte sécurité, Non-conformité, Anomalie, À corriger, Défaut constaté, Réserve, À surveiller, Observation, Point d'amélioration, Remarque
NIVEAUX : CRITIQUE=Alerte sécurité | ÉLEVÉ=Non-conformité/Anomalie/À corriger | MOYEN=Réserve/Surveiller/Défaut | FAIBLE=Observation/Amélioration | MINIMAL=Remarque
STATUTS : CRITIQUE→"À faire" | ÉLEVÉ→"À planifier" | MOYEN→"À planifier" | FAIBLE→"À valider" | MINIMAL→"À valider"

Réponds UNIQUEMENT en JSON valide sans texte avant/après :
{
  "metadata": {"date_intervention":"","societe":"","client":"","site":"","materiel":"","etat":""},
  "problemes": [{"num":1,"classification":"","description":"","action_corrective":"","niveau_risque":"","statut":"","reference_reglementaire":""}],
  "statistiques": {"total":0,"critique":0,"eleve":0,"moyen":0,"faible":0,"minimal":0},
  "resume_executif": ""
}

Texte du rapport :
{texte}
"""


def _extraire_texte_pdf(chemin):
    """Extrait le texte d'un PDF via pdfplumber."""
    try:
        import pdfplumber
        texte = ""
        with pdfplumber.open(chemin) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texte += t + "\n"
        return texte.strip()
    except Exception as e:
        return ""


def _parse_gemini_json(texte_rep):
    """Parse robuste du JSON retourne par Gemini sans recursion."""
    import json as _json, re as _re
    texte_rep = _re.sub(r'`' + '`' + '`' + r'json\s*', '', texte_rep)
    texte_rep = _re.sub(r'`' + '`' + '`' + r'\s*', '', texte_rep)
    texte_rep = texte_rep.strip()
    # 1. Essai direct
    try:
        return _json.loads(texte_rep)
    except Exception:
        pass
    # 2. Extraction par indices (evite regex DOTALL -> max recursion)
    s = texte_rep.find('{')
    e = texte_rep.rfind('}')
    if s != -1 and e != -1 and e > s:
        candidate = texte_rep[s:e + 1]
        try:
            return _json.loads(candidate)
        except Exception:
            pass
        # 3. Remplacer sauts de ligne par espace
        cleaned = candidate
        for ch in [chr(13)+chr(10), chr(13), chr(10)]:
            cleaned = cleaned.replace(ch, ' ')
        try:
            return _json.loads(cleaned)
        except Exception:
            pass
        # 4. JSON tronque: fermer
        try:
            fixed = cleaned.rstrip().rstrip(',')
            n_quotes = fixed.count(chr(34)) - fixed.count(chr(92)+chr(34))
            if n_quotes % 2 == 1:
                fixed += chr(34)
            if not fixed.endswith('}'):
                fixed += '}'
            return _json.loads(fixed)
        except Exception:
            pass
    raise ValueError('JSON Gemini non parseable: ' + texte_rep[:200])


def _analyser_gemini(texte, api_key):
    """Envoie le texte à Gemini et retourne le JSON parsé."""
    import urllib.request, json as _json, re as _re
    prompt = PROMPT_INSPECTION.replace("{texte}", texte[:40000])
    body = _json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 8192}
    }).encode("utf-8")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=120) as r:
        result = _json.loads(r.read().decode("utf-8"))
    texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
    texte_rep = _re.sub(r"```json\s*", "", texte_rep)
    texte_rep = _re.sub(r"```\s*", "", texte_rep)
    return _parse_gemini_json(texte_rep)


@app.route('/api/analyse-pdf/config', methods=['GET'])
@login_required
def get_analyse_config():
    has_key = bool(app.config.get('GEMINI_API_KEY'))
    return jsonify({'has_server_key': has_key})


@app.route('/api/analyse-pdf/upload', methods=['POST'])
@login_required
def upload_analyse_pdf():
    import os as _os, tempfile as _tmp, json as _json
    from datetime import date as _date

    etab = _get_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Seuls les PDF sont acceptés'}), 400

    # Priorité : clé serveur (Railway env var) > clé fournie par l'utilisateur
    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API Gemini requise — configurez GEMINI_API_KEY sur Railway ou saisissez-la dans l\'interface'}), 400

    # Sauvegarder temporairement
    with _tmp.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        texte = _extraire_texte_pdf(tmp_path)

        if texte:
            # PDF textuel : envoi du texte à Gemini
            data = _analyser_gemini(texte, api_key)
        else:
            # PDF scanné (images) : envoi direct en base64 à Gemini vision
            import base64 as _b64, urllib.request as _req, json as _json2, re as _re2
            with open(tmp_path, 'rb') as _pf:
                pdf_b64 = _b64.b64encode(_pf.read()).decode('utf-8')
            prompt_vision = PROMPT_INSPECTION.replace("{texte}", "[PDF envoyé en image — lis le contenu du document]")
            # Remplacer la partie texte par une instruction pour lire le PDF visuel
            prompt_vision = (
                "Tu es un expert en sécurité bâtiment et réglementation ERP. "
                "Analyse ce rapport d'inspection (PDF scanné fourni en image).\n\n"
                "CLASSIFICATIONS : Alerte sécurité, Non-conformité, Anomalie, À corriger, Défaut constaté, Réserve, À surveiller, Observation, Point d'amélioration, Remarque\n"
                "NIVEAUX : CRITIQUE=Alerte sécurité | ÉLEVÉ=Non-conformité/Anomalie/À corriger | MOYEN=Réserve/Surveiller/Défaut | FAIBLE=Observation/Amélioration | MINIMAL=Remarque\n"
                "STATUTS : CRITIQUE→\"À faire\" | ÉLEVÉ→\"À planifier\" | MOYEN→\"À planifier\" | FAIBLE→\"À valider\" | MINIMAL→\"À valider\"\n\n"
                "Réponds UNIQUEMENT en JSON valide sans texte avant/après :\n"
                '{"metadata": {"date_intervention":"","societe":"","client":"","site":"","materiel":"","etat":""},'
                '"problemes": [{"num":1,"classification":"","description":"","action_corrective":"","niveau_risque":"","statut":"","reference_reglementaire":""}],'
                '"statistiques": {"total":0,"critique":0,"eleve":0,"moyen":0,"faible":0,"minimal":0},'
                '"resume_executif": ""}'
            )
            body_v = _json2.dumps({
                "contents": [{"parts": [
                    {"text": prompt_vision},
                    {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 8192}
            }).encode("utf-8")
            url_v = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            req_v = _req.Request(url_v, data=body_v, headers={"Content-Type": "application/json"})
            with _req.urlopen(req_v, timeout=120) as r_v:
                result_v = _json2.loads(r_v.read().decode("utf-8"))
            texte_rep = result_v["candidates"][0]["content"]["parts"][0]["text"].strip()
            texte_rep = _re2.sub(r"```json\s*", "", texte_rep)
            texte_rep = _re2.sub(r"```\s*", "", texte_rep)
            data = _json2.loads(texte_rep)

        meta  = data.get("metadata", {})
        stats = data.get("statistiques", {})

        analyse = AnalysePDF(
            etab_id=etab.id,
            date_analyse=_date.today().isoformat(),
            nom_fichier=f.filename,
            societe=meta.get("societe", ""),
            client=meta.get("client", ""),
            site=meta.get("site", ""),
            materiel=meta.get("materiel", ""),
            etat=meta.get("etat", ""),
            total_problemes=stats.get("total", 0),
            nb_critique=stats.get("critique", 0),
            nb_eleve=stats.get("eleve", 0),
            nb_moyen=stats.get("moyen", 0),
            nb_faible=stats.get("faible", 0),
            resume_executif=data.get("resume_executif", ""),
            data_json=_json.dumps(data, ensure_ascii=False)
        )
        db.session.add(analyse)
        db.session.commit()

        d = analyse.to_dict()
        d['data'] = data
        return jsonify(d), 201

    except Exception as e:
        return jsonify({'error': f'Erreur analyse : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


PROMPT_CONTRAT = """
Tu es un expert en gestion de contrats de prestataires. Analyse ce contrat et extrais les informations clés.

Réponds UNIQUEMENT en JSON valide sans texte avant/après :
{
  "nom": "nom du prestataire / société",
  "service": "type de service parmi : Nettoyage, Maintenance, Sécurité, Restauration, Espaces verts, Fournitures, Informatique, Téléphonie, Transport / Logistique, Comptabilité / Finance, Juridique, Formation, Communication / Marketing, Déchets / Recyclage, Énergie, Assurance, Désinsectisation, Autre",
  "contact": "nom du contact principal",
  "email": "email de contact",
  "telephone": "numéro de téléphone",
  "date_debut": "date de début du contrat au format YYYY-MM-DD",
  "date_fin": "date de fin du contrat au format YYYY-MM-DD",
  "montant": 0,
  "frequence": "fréquence de paiement : Mensuel, Trimestriel, Annuel, Ponctuel",
  "prestations": "description courte des prestations",
  "notes": "informations importantes supplémentaires"
}

Si une information est absente, laisse la valeur vide ("") ou 0 pour le montant.

Texte du contrat :
{texte}
"""


@app.route('/api/prestataires/scan-contrat', methods=['POST'])
@login_required
def scan_contrat_ia():
    import os as _os, tempfile as _tmp, json as _json, re as _re, urllib.request
    from datetime import date as _date

    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Seuls les PDF sont acceptés'}), 400

    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API Gemini requise'}), 400

    with _tmp.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        import base64 as _b64
        texte = _extraire_texte_pdf(tmp_path)
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"

        if texte:
            # PDF textuel
            prompt = PROMPT_CONTRAT.replace("{texte}", texte[:40000])
            body = _json.dumps({
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")
        else:
            # PDF scanné → Vision Gemini
            pdf_b64 = _b64.b64encode(open(tmp_path, 'rb').read()).decode('utf-8')
            prompt_vision = PROMPT_CONTRAT.replace("{texte}", "[Analyse le contrat visible dans ce PDF scanné]")
            body = _json.dumps({
                "contents": [{"parts": [
                    {"text": prompt_vision},
                    {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")

        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=90) as r:
            result = _json.loads(r.read().decode("utf-8"))
        texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
        texte_rep = _re.sub(r"```json\s*", "", texte_rep)
        texte_rep = _re.sub(r"```\s*", "", texte_rep)
        data = _parse_gemini_json(texte_rep)
        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': f'Erreur scan : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.route('/api/vehicules/scan-document', methods=['POST'])
@login_required
def scan_document_vehicule():
    """Scan carte grise, CT ou assurance — extrait les infos du vehicule."""
    import os as _os, tempfile as _tmp, json as _json, re as _re, urllib.request, base64 as _b64

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    fname = f.filename.lower()
    if not (fname.endswith('.pdf') or fname.endswith('.jpg') or fname.endswith('.jpeg') or fname.endswith('.png')):
        return jsonify({'error': 'Format accepté : PDF, JPG, PNG'}), 400

    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API requise'}), 400

    suffix = '.' + fname.rsplit('.', 1)[-1]
    with _tmp.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    prompt = (
        "Analyse ce document de véhicule (carte grise, certificat de contrôle technique CT, "
        "ou attestation d'assurance). Extrais les informations en JSON strict, sans markdown :\n"
        '{"immatriculation": "numéro immatriculation (champ A sur carte grise, ex: AB-123-CD)", '
        '"marque": "marque du véhicule (champ D.1 sur carte grise)", '
        '"modele": "modèle / version (champ D.2 ou D.3)", '
        '"annee": "année de première mise en circulation (champ B, format YYYY)", '
        '"date_ct": "date limite du prochain CT au format JJ-MM-AAAA — cherche validité, prochain CT, date limite sur le CT", '
        '"date_assurance": "date de fin de validité de l\'assurance au format JJ-MM-AAAA", '
        '"remarques_ct": "liste des défauts notés sur le rapport CT séparés par ; — vide si aucun ou si ce n\'est pas un CT"}\n'
        "Règles : si une info est absente laisse la valeur vide. "
        "Retourne UNIQUEMENT le JSON, rien d'autre."
    )

    try:
        if suffix == '.pdf':
            texte = _extraire_texte_pdf(tmp_path)
            if texte:
                body = _json.dumps({
                    "contents": [{"parts": [{"text": prompt + "\n\nTexte du document :\n" + texte[:40000]}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
            else:
                pdf_b64 = _b64.b64encode(open(tmp_path, 'rb').read()).decode('utf-8')
                body = _json.dumps({
                    "contents": [{"parts": [
                        {"text": prompt + "\n\n[Document scanné en image]"},
                        {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                    ]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
        else:
            with open(tmp_path, 'rb') as img_f:
                img_data = _b64.b64encode(img_f.read()).decode('utf-8')
            mime = 'image/jpeg' if suffix in ('.jpg', '.jpeg') else 'image/png'
            body = _json.dumps({
                "contents": [{"parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime, "data": img_data}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=90) as r:
            result = _json.loads(r.read().decode("utf-8"))
        texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
        texte_rep = _re.sub(r"```json\s*", "", texte_rep)
        texte_rep = _re.sub(r"```\s*", "", texte_rep)
        data = _parse_gemini_json(texte_rep)
        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': f'Erreur scan : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.route('/api/analyse-pdf/analyses', methods=['GET'])
@app.route('/api/analyse-pdf/historique', methods=['GET'])
@login_required
def get_historique_analyses():
    etab = _get_etab()
    if not etab:
        return jsonify([])
    analyses = AnalysePDF.query.filter_by(etab_id=etab.id).order_by(AnalysePDF.created_at.desc()).all()
    return jsonify([a.to_dict() for a in analyses])


@app.route('/api/analyse-pdf/<int:aid>', methods=['GET'])
@login_required
def get_analyse_detail(aid):
    etab = _get_etab()
    a = AnalysePDF.query.filter_by(id=aid, etab_id=etab.id).first_or_404()
    d = a.to_dict()
    import json as _json
    try:
        d['data'] = _json.loads(a.data_json)
    except Exception:
        d['data'] = {}
    return jsonify(d)


@app.route('/api/analyse-pdf/<int:aid>', methods=['DELETE'])
@login_required
def delete_analyse(aid):
    etab = _get_etab()
    a = AnalysePDF.query.filter_by(id=aid, etab_id=etab.id).first_or_404()
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  PARC AUTO
# ══════════════════════════════════════════════════════

def _get_etab():
    etab = Etablissement.query.filter_by(user_id=current_user.id, is_current=True).first()
    if not etab:
        etab = Etablissement.query.filter_by(user_id=current_user.id).first()
    return etab

def _mois_restants(date_str):
    """Retourne le nombre de mois entre aujourd'hui et date_str (format JJ-MM-AAAA ou AAAA-MM-JJ)."""
    if not date_str:
        return None
    from datetime import date as _date
    try:
        if '-' in date_str:
            parts = date_str.split('-')
            if len(parts[0]) == 4:
                d = _date(int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                d = _date(int(parts[2]), int(parts[1]), int(parts[0]))
        else:
            return None
        today = _date.today()
        delta_days = (d - today).days
        return round(delta_days / 30.44)
    except Exception:
        return None


@app.route('/api/vehicules', methods=['GET'])
@login_required
def get_vehicules():
    etab = _get_etab()
    if not etab:
        return jsonify([])
    vehicules = Vehicule.query.filter_by(etab_id=etab.id).order_by(Vehicule.marque).all()
    result = []
    for v in vehicules:
        d = v.to_dict()
        d['mois_ct'] = _mois_restants(v.date_ct)
        d['mois_assurance'] = _mois_restants(v.date_assurance)
        result.append(d)
    return jsonify(result)


@app.route('/api/vehicules', methods=['POST'])
@login_required
def add_vehicule():
    etab = _get_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    immat = (data.get('immatriculation') or '').strip().upper()
    if not immat:
        return jsonify({'error': 'Immatriculation requise'}), 400
    if Vehicule.query.filter_by(etab_id=etab.id, immatriculation=immat).first():
        return jsonify({'error': 'Immatriculation deja existante'}), 409
    v = Vehicule(
        etab_id=etab.id,
        immatriculation=immat,
        marque=data.get('marque', ''),
        modele=data.get('modele', ''),
        annee=int(data.get('annee') or 0),
        kilometrage=int(data.get('kilometrage') or 0),
        couleur=data.get('couleur', '#808080'),
        conducteur=data.get('conducteur', ''),
        date_ct=data.get('date_ct', ''),
        date_assurance=data.get('date_assurance', ''),
        remarques_ct=data.get('remarques_ct', ''),
        notes=data.get('notes', '')
    )
    db.session.add(v)
    db.session.commit()
    d = v.to_dict()
    d['mois_ct'] = _mois_restants(v.date_ct)
    d['mois_assurance'] = _mois_restants(v.date_assurance)
    return jsonify(d), 201


@app.route('/api/vehicules/<int:vid>', methods=['PUT'])
@login_required
def update_vehicule(vid):
    etab = _get_etab()
    v = Vehicule.query.filter_by(id=vid, etab_id=etab.id).first_or_404()
    data = request.get_json()
    for field in ['immatriculation', 'marque', 'modele', 'conducteur', 'couleur',
                  'date_ct', 'date_assurance', 'remarques_ct', 'notes']:
        if field in data:
            setattr(v, field, data[field])
    if 'annee' in data:
        v.annee = int(data['annee'] or 0)
    if 'kilometrage' in data:
        v.kilometrage = int(data['kilometrage'] or 0)
    db.session.commit()
    d = v.to_dict()
    d['mois_ct'] = _mois_restants(v.date_ct)
    d['mois_assurance'] = _mois_restants(v.date_assurance)
    return jsonify(d)


@app.route('/api/vehicules/<int:vid>', methods=['DELETE'])
@login_required
def delete_vehicule(vid):
    etab = _get_etab()
    v = Vehicule.query.filter_by(id=vid, etab_id=etab.id).first_or_404()
    db.session.delete(v)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/vehicules/<int:vid>/km', methods=['POST'])
@login_required
def update_km(vid):
    etab = _get_etab()
    v = Vehicule.query.filter_by(id=vid, etab_id=etab.id).first_or_404()
    data = request.get_json()
    v.kilometrage = int(data.get('kilometrage') or v.kilometrage)
    db.session.commit()
    return jsonify({'ok': True, 'kilometrage': v.kilometrage})


@app.route('/api/vehicules/alertes', methods=['GET'])
@login_required
def get_alertes_vehicules():
    etab = _get_etab()
    if not etab:
        return jsonify([])
    vehicules = Vehicule.query.filter_by(etab_id=etab.id).all()
    alertes = []
    for v in vehicules:
        mois_ct  = _mois_restants(v.date_ct)
        mois_ass = _mois_restants(v.date_assurance)
        alerte_ct  = mois_ct  is not None and mois_ct  <= 3
        alerte_ass = mois_ass is not None and mois_ass <= 3
        if alerte_ct or alerte_ass:
            alertes.append({
                'id': v.id, 'immatriculation': v.immatriculation,
                'marque': v.marque, 'modele': v.modele,
                'date_ct': v.date_ct, 'mois_ct': mois_ct, 'alerte_ct': alerte_ct,
                'date_assurance': v.date_assurance, 'mois_assurance': mois_ass, 'alerte_ass': alerte_ass
            })
    return jsonify(alertes)


# Entretiens

@app.route('/api/entretiens', methods=['GET'])
@login_required
def get_entretiens():
    etab = _get_etab()
    vid = request.args.get('vehicule_id', type=int)
    q = Entretien.query.join(Vehicule).filter(Vehicule.etab_id == etab.id)
    if vid:
        q = q.filter(Entretien.vehicule_id == vid)
    return jsonify([e.to_dict() for e in q.order_by(Entretien.date.desc()).all()])


@app.route('/api/entretiens', methods=['POST'])
@login_required
def add_entretien():
    etab = _get_etab()
    data = request.get_json()
    vid = int(data.get('vehicule_id') or 0)
    v = Vehicule.query.filter_by(id=vid, etab_id=etab.id).first_or_404()
    e = Entretien(
        vehicule_id=v.id,
        date=data.get('date', ''),
        type_entretien=data.get('type_entretien', ''),
        kilometrage=int(data.get('kilometrage') or 0),
        description=data.get('description', ''),
        cout=float(data.get('cout') or 0)
    )
    db.session.add(e)
    db.session.commit()
    return jsonify(e.to_dict()), 201


@app.route('/api/entretiens/<int:eid>', methods=['DELETE'])
@login_required
def delete_entretien(eid):
    etab = _get_etab()
    e = Entretien.query.join(Vehicule).filter(Vehicule.etab_id == etab.id, Entretien.id == eid).first_or_404()
    db.session.delete(e)
    db.session.commit()
    return jsonify({'ok': True})


# Carburant

@app.route('/api/carburant', methods=['GET'])
@login_required
def get_carburant():
    etab = _get_etab()
    vid = request.args.get('vehicule_id', type=int)
    q = Carburant.query.join(Vehicule).filter(Vehicule.etab_id == etab.id)
    if vid:
        q = q.filter(Carburant.vehicule_id == vid)
    return jsonify([c.to_dict() for c in q.order_by(Carburant.date.desc()).all()])


@app.route('/api/carburant', methods=['POST'])
@login_required
def add_carburant():
    etab = _get_etab()
    data = request.get_json()
    vid = int(data.get('vehicule_id') or 0)
    v = Vehicule.query.filter_by(id=vid, etab_id=etab.id).first_or_404()
    c = Carburant(
        vehicule_id=v.id,
        date=data.get('date', ''),
        litres=float(data.get('litres') or 0),
        cout=float(data.get('cout') or 0),
        kilometrage=int(data.get('kilometrage') or 0)
    )
    db.session.add(c)
    # Mettre à jour km si plus élevé
    if c.kilometrage > v.kilometrage:
        v.kilometrage = c.kilometrage
    db.session.commit()
    return jsonify(c.to_dict()), 201


@app.route('/api/carburant/<int:cid>', methods=['DELETE'])
@login_required
def delete_carburant(cid):
    etab = _get_etab()
    c = Carburant.query.join(Vehicule).filter(Vehicule.etab_id == etab.id, Carburant.id == cid).first_or_404()
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  UTILITAIRES
# ══════════════════════════════════════════════════════

def data_today():
    from datetime import date
    return date.today().isoformat()


# ══════════════════════════════════════════════════════
#  SCAN FICHE STOCK
# ══════════════════════════════════════════════════════

@app.route('/api/stock/scan-fiche', methods=['POST'])
@login_required
def scan_fiche_stock():
    """Scan une fiche de distribution produits — extrait date, departement, produits+quantites."""
    import os as _os, tempfile as _tmp, json as _json, re as _re, urllib.request, base64 as _b64
    import datetime as _dt

    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    fname = f.filename.lower()
    if not (fname.endswith('.pdf') or fname.endswith('.jpg') or fname.endswith('.jpeg') or fname.endswith('.png')):
        return jsonify({'error': 'Format accepté : PDF, JPG, PNG'}), 400

    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API requise'}), 400

    # Catalogue produits pour le prompt
    produits = StockProduit.query.filter_by(etab_id=etab.id).all()
    product_lines = '\n'.join(f"- {p.id} | {p.nom} ({p.unite})" for p in produits) or '- aucun'
    today = _dt.date.today().strftime('%d-%m-%Y')

    prompt = (
        "Tu analyses une fiche de distribution de produits (nettoyage, entretien, etc.).\n"
        "Le document peut être manuscrit, imprimé ou mixte.\n\n"
        "Objectif :\n"
        "- Lire la date (manuscrite ou imprimée)\n"
        "- Lire le nom du département/groupe/service\n"
        "- Pour chaque produit : lire le nom et la quantité REMISE (dernière colonne, ignore la quantité demandée)\n"
        "- Ne garder que les lignes avec une quantité remise > 0\n\n"
        "Renvoie UNIQUEMENT ce JSON valide :\n"
        '{"date": "DD-MM-YYYY", "departement": "nom du groupe", '
        '"entries": [{"produit_id": 123, "produit_nom": "nom exact", "quantite": 2}]}\n\n'
        f"Règles : date au format DD-MM-YYYY (si illisible utilise {today}). "
        "produit_id doit venir du catalogue ci-dessous ou null si absent. "
        "Ne renvoie pas de texte hors JSON.\n\n"
        f"Catalogue produits :\n{product_lines}"
    )

    suffix = '.' + fname.rsplit('.', 1)[-1]
    with _tmp.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        if suffix == '.pdf':
            texte = _extraire_texte_pdf(tmp_path)
            if texte:
                body = _json.dumps({
                    "contents": [{"parts": [{"text": prompt + "\n\nTexte du document :\n" + texte[:40000]}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
            else:
                pdf_b64 = _b64.b64encode(open(tmp_path, 'rb').read()).decode('utf-8')
                body = _json.dumps({
                    "contents": [{"parts": [
                        {"text": prompt + "\n\n[Document scanné]"},
                        {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                    ]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
        else:
            with open(tmp_path, 'rb') as img_f:
                img_data = _b64.b64encode(img_f.read()).decode('utf-8')
            mime = 'image/jpeg' if suffix in ('.jpg', '.jpeg') else 'image/png'
            body = _json.dumps({
                "contents": [{"parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime, "data": img_data}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=90) as r:
            result = _json.loads(r.read().decode("utf-8"))
        texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
        texte_rep = _re.sub(r"```json\s*", "", texte_rep)
        texte_rep = _re.sub(r"```\s*", "", texte_rep)
        data = _parse_gemini_json(texte_rep)
        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': f'Erreur scan : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


# ══════════════════════════════════════════════════════
#  API STOCK
# ══════════════════════════════════════════════════════

@app.route('/api/stock/produits', methods=['GET'])
@login_required
def api_stock_produits():
    etab = get_current_etab()
    if not etab: return jsonify([])
    q = request.args.get('q', '').lower()
    cat = request.args.get('categorie', '')
    prods = StockProduit.query.filter_by(etab_id=etab.id).order_by(StockProduit.nom).all()
    result = [p.to_dict() for p in prods
              if (not q or q in p.nom.lower())
              and (not cat or p.categorie == cat)]
    return jsonify(result)


@app.route('/api/stock/produits', methods=['POST'])
@login_required
def api_stock_add_produit():
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    nom = (data.get('nom') or '').strip()
    if not nom: return jsonify({'error': 'Nom requis'}), 400
    p = StockProduit(
        etab_id=etab.id, nom=nom,
        categorie=data.get('categorie', ''),
        quantite=float(data.get('quantite', 0)),
        unite=data.get('unite', 'unité'),
        seuil_alerte=float(data.get('seuil_alerte', 0)),
        emplacement=data.get('emplacement', '')
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@app.route('/api/stock/produits/<int:pid>', methods=['PUT'])
@login_required
def api_stock_update_produit(pid):
    etab = get_current_etab()
    p = StockProduit.query.filter_by(id=pid, etab_id=etab.id).first_or_404()
    data = request.get_json()
    p.nom = data.get('nom', p.nom)
    p.categorie = data.get('categorie', p.categorie)
    p.unite = data.get('unite', p.unite)
    p.seuil_alerte = float(data.get('seuil_alerte', p.seuil_alerte))
    p.emplacement = data.get('emplacement', p.emplacement)
    db.session.commit()
    return jsonify(p.to_dict())


@app.route('/api/stock/produits/<int:pid>', methods=['DELETE'])
@login_required
def api_stock_delete_produit(pid):
    etab = get_current_etab()
    p = StockProduit.query.filter_by(id=pid, etab_id=etab.id).first_or_404()
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/stock/mouvements', methods=['GET'])
@login_required
def api_stock_mouvements():
    etab = get_current_etab()
    if not etab: return jsonify([])
    type_filter = request.args.get('type', '')
    jours = int(request.args.get('jours', 0))
    query = (StockMouvement.query
             .join(StockProduit, StockMouvement.produit_id == StockProduit.id)
             .filter(StockProduit.etab_id == etab.id))
    if type_filter:
        query = query.filter(StockMouvement.type == type_filter)
    if jours > 0:
        from datetime import date, timedelta
        depuis = (date.today() - timedelta(days=jours)).isoformat()
        query = query.filter(StockMouvement.date >= depuis)
    mvts = query.order_by(StockMouvement.date.desc(), StockMouvement.id.desc()).limit(500).all()
    return jsonify([m.to_dict() for m in mvts])


@app.route('/api/stock/mouvements', methods=['POST'])
@login_required
def api_stock_add_mouvement():
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    pid = data.get('produit_id')
    typ = data.get('type', 'sortie')
    qte = float(data.get('quantite', 0))
    if not pid or qte <= 0:
        return jsonify({'error': 'produit_id et quantite requis'}), 400

    p = StockProduit.query.filter_by(id=pid, etab_id=etab.id).first()
    if not p: return jsonify({'error': 'Produit introuvable'}), 404

    if typ == 'sortie' and p.quantite < qte:
        return jsonify({'error': f'Stock insuffisant ({p.quantite} {p.unite} disponibles)'}), 409

    from datetime import date
    m = StockMouvement(
        produit_id=pid, type=typ, quantite=qte,
        personne=data.get('personne', ''),
        departement=data.get('departement', ''),
        date=data.get('date', date.today().isoformat()),
        notes=data.get('notes', '')
    )
    db.session.add(m)
    if typ == 'sortie':
        p.quantite -= qte
    else:
        p.quantite += qte
    db.session.commit()
    return jsonify({'ok': True, 'nouvelle_quantite': p.quantite}), 201


@app.route('/api/stock/alertes', methods=['GET'])
@login_required
def api_stock_alertes():
    etab = get_current_etab()
    if not etab: return jsonify([])
    prods = StockProduit.query.filter_by(etab_id=etab.id).all()
    alertes = [p.to_dict() for p in prods if p.seuil_alerte > 0 and p.quantite <= p.seuil_alerte]
    return jsonify(alertes)


@app.route('/api/stock/categories', methods=['GET'])
@login_required
def api_stock_categories():
    etab = get_current_etab()
    if not etab: return jsonify([])
    cats = db.session.query(StockProduit.categorie).filter(
        StockProduit.etab_id == etab.id,
        StockProduit.categorie != ''
    ).distinct().all()
    return jsonify(sorted([c[0] for c in cats]))


@app.route('/api/stock/stats', methods=['GET'])
@login_required
def api_stock_stats():
    etab = get_current_etab()
    if not etab: return jsonify({})
    prods = StockProduit.query.filter_by(etab_id=etab.id).all()
    total = len(prods)
    alertes = sum(1 for p in prods if p.seuil_alerte > 0 and p.quantite <= p.seuil_alerte)
    ruptures = sum(1 for p in prods if p.quantite == 0)
    return jsonify({'total_produits': total, 'alertes': alertes, 'ruptures': ruptures})


# ══════════════════════════════════════════════════════
#  API CLÉS
# ══════════════════════════════════════════════════════

def _log_cle_event(etab_id, event_type, details):
    from datetime import datetime
    h = HistoriqueCle(etab_id=etab_id, event_type=event_type, details=details,
                      event_date=datetime.now().strftime('%d/%m/%Y %H:%M'))
    db.session.add(h)


@app.route('/api/cles/scan-fiche', methods=['POST'])
@login_required
def scan_fiche_cles():
    """Scan une fiche de distribution de cles — extrait employe, date, cles attribuees."""
    import os as _os, tempfile as _tmp, json as _json, re as _re, urllib.request, base64 as _b64
    import datetime as _dt, unicodedata as _ud

    etab = get_current_etab()
    if not etab:
        return jsonify({'error': 'Pas d\'etablissement'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    f = request.files['file']
    fname = f.filename.lower()
    if not (fname.endswith('.pdf') or fname.endswith('.jpg') or fname.endswith('.jpeg') or fname.endswith('.png')):
        return jsonify({'error': 'Format accepté : PDF, JPG, PNG'}), 400

    api_key = app.config.get('GEMINI_API_KEY') or request.form.get('api_key', '').strip()
    if not api_key:
        return jsonify({'error': 'Clé API requise'}), 400

    # Catalogues pour le prompt
    employes = EmployeCle.query.filter_by(etab_id=etab.id).all()
    cles = CleItem.query.filter_by(etab_id=etab.id).all()
    emp_lines = '\n'.join(f"- {e.id} | {e.nom} {e.prenom} | {e.poste or ''}" for e in employes) or '- aucun'
    cle_lines = '\n'.join(f"- {c.id} | num={c.numero or '-'} | {c.nom}" for c in cles) or '- aucune'
    today = _dt.date.today().strftime('%d-%m-%Y')

    prompt = (
        "Tu analyses une fiche de distribution de clés (manuscrite ou photographiée).\n\n"
        "Objectif :\n"
        "- Reconnaître l'employé concerné\n"
        "- Reconnaître la date d'attribution\n"
        "- Reconnaître chaque clé notée sur la fiche\n\n"
        "Renvoie UNIQUEMENT ce JSON valide :\n"
        '{"employee_id": 123, "employee_name_raw": "nom lu", "assignment_date": "DD-MM-YYYY", '
        '"notes": "commentaire", "entries": [{"raw": "texte vu", "label": "interprétation", '
        '"key_id": 456, "confidence": "high"}]}\n\n'
        f"Règles : employee_id et key_id doivent venir des catalogues ou null. "
        f"date au format DD-MM-YYYY (si illisible utilise {today}). "
        "confidence = high/medium/low. Ne renvoie pas de texte hors JSON.\n\n"
        f"Catalogue employés :\n{emp_lines}\n\n"
        f"Catalogue clés :\n{cle_lines}"
    )

    suffix = '.' + fname.rsplit('.', 1)[-1]
    with _tmp.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        if suffix == '.pdf':
            texte = _extraire_texte_pdf(tmp_path)
            if texte:
                body = _json.dumps({
                    "contents": [{"parts": [{"text": prompt + "\n\nTexte du document :\n" + texte[:40000]}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
            else:
                pdf_b64 = _b64.b64encode(open(tmp_path, 'rb').read()).decode('utf-8')
                body = _json.dumps({
                    "contents": [{"parts": [
                        {"text": prompt + "\n\n[Document scanné]"},
                        {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
                    ]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
                }).encode("utf-8")
        else:
            with open(tmp_path, 'rb') as img_f:
                img_data = _b64.b64encode(img_f.read()).decode('utf-8')
            mime = 'image/jpeg' if suffix in ('.jpg', '.jpeg') else 'image/png'
            body = _json.dumps({
                "contents": [{"parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime, "data": img_data}}
                ]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
            }).encode("utf-8")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=90) as r:
            result = _json.loads(r.read().decode("utf-8"))
        texte_rep = result["candidates"][0]["content"]["parts"][0]["text"].strip()
        texte_rep = _re.sub(r"```json\s*", "", texte_rep)
        texte_rep = _re.sub(r"```\s*", "", texte_rep)
        data = _parse_gemini_json(texte_rep)
        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': f'Erreur scan : {str(e)}'}), 500
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.route('/api/cles/items', methods=['GET'])
@app.route('/api/cles/cles', methods=['GET'])
@login_required
def api_cles_list():
    etab = get_current_etab()
    if not etab: return jsonify([])
    items = CleItem.query.filter_by(etab_id=etab.id).order_by(CleItem.nom).all()
    result = []
    for c in items:
        d = c.to_dict()
        attribuees = AttributionCle.query.filter_by(cle_id=c.id).filter(
            (AttributionCle.date_retour == '') | (AttributionCle.date_retour == None)
        ).count()
        d['attribuees'] = attribuees
        d['disponibles'] = max(0, c.quantite_totale - attribuees)
        result.append(d)
    return jsonify(result)


@app.route('/api/cles/cles', methods=['POST'])
@login_required
def api_cles_add():
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    nom = (data.get('nom') or '').strip()
    if not nom: return jsonify({'error': 'Nom requis'}), 400
    c = CleItem(etab_id=etab.id, nom=nom,
                numero=data.get('numero', ''),
                quantite_totale=int(data.get('quantite_totale', 1)))
    db.session.add(c)
    db.session.flush()
    _log_cle_event(etab.id, 'cle_ajoutee', f'Clé ajoutée: {nom} (n°{c.numero})')
    db.session.commit()
    return jsonify(c.to_dict()), 201


@app.route('/api/cles/cles/<int:cle_id>', methods=['PUT'])
@login_required
def api_cles_update(cle_id):
    etab = get_current_etab()
    c = CleItem.query.filter_by(id=cle_id, etab_id=etab.id).first_or_404()
    data = request.get_json()
    c.nom = data.get('nom', c.nom)
    c.numero = data.get('numero', c.numero)
    c.quantite_totale = int(data.get('quantite_totale', c.quantite_totale))
    _log_cle_event(etab.id, 'cle_modifiee', f'Clé modifiée: {c.nom}')
    db.session.commit()
    return jsonify(c.to_dict())


@app.route('/api/cles/cles/<int:cle_id>', methods=['DELETE'])
@login_required
def api_cles_delete(cle_id):
    etab = get_current_etab()
    c = CleItem.query.filter_by(id=cle_id, etab_id=etab.id).first_or_404()
    _log_cle_event(etab.id, 'cle_supprimee', f'Clé supprimée: {c.nom}')
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/cles/employes', methods=['GET'])
@login_required
def api_employes_list():
    etab = get_current_etab()
    if not etab: return jsonify([])
    employes = EmployeCle.query.filter_by(etab_id=etab.id).order_by(EmployeCle.nom).all()
    result = []
    for e in employes:
        d = e.to_dict()
        cles_actives = AttributionCle.query.filter_by(employe_id=e.id).filter(
            (AttributionCle.date_retour == '') | (AttributionCle.date_retour == None)
        ).all()
        d['cles'] = [{'id': a.cle_id, 'nom': a.cle.nom if a.cle else '', 'numero': a.cle.numero if a.cle else '',
                      'attribution_id': a.id, 'date_attribution': a.date_attribution} for a in cles_actives]
        d['nb_cles'] = len(cles_actives)
        result.append(d)
    return jsonify(result)


@app.route('/api/cles/employes', methods=['POST'])
@login_required
def api_employes_add():
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    nom = (data.get('nom') or '').strip()
    if not nom: return jsonify({'error': 'Nom requis'}), 400
    e = EmployeCle(etab_id=etab.id, nom=nom,
                   prenom=data.get('prenom', ''),
                   type_contrat=data.get('type_contrat', 'CDI'),
                   poste=data.get('poste', ''),
                   date_arrivee=data.get('date_arrivee', ''),
                   date_depart=data.get('date_depart', ''))
    db.session.add(e)
    db.session.flush()
    _log_cle_event(etab.id, 'employe_ajoute', f'Employé ajouté: {e.prenom} {e.nom}')
    db.session.commit()
    return jsonify(e.to_dict()), 201


@app.route('/api/cles/employes/<int:emp_id>', methods=['PUT'])
@login_required
def api_employes_update(emp_id):
    etab = get_current_etab()
    e = EmployeCle.query.filter_by(id=emp_id, etab_id=etab.id).first_or_404()
    data = request.get_json()
    e.nom = data.get('nom', e.nom)
    e.prenom = data.get('prenom', e.prenom)
    e.type_contrat = data.get('type_contrat', e.type_contrat)
    e.poste = data.get('poste', e.poste)
    e.date_arrivee = data.get('date_arrivee', e.date_arrivee)
    e.date_depart = data.get('date_depart', e.date_depart)
    _log_cle_event(etab.id, 'employe_modifie', f'Employé modifié: {e.prenom} {e.nom}')
    db.session.commit()
    return jsonify(e.to_dict())


@app.route('/api/cles/employes/<int:emp_id>', methods=['DELETE'])
@login_required
def api_employes_delete(emp_id):
    etab = get_current_etab()
    e = EmployeCle.query.filter_by(id=emp_id, etab_id=etab.id).first_or_404()
    _log_cle_event(etab.id, 'employe_supprime', f'Employé supprimé: {e.prenom} {e.nom}')
    db.session.delete(e)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/cles/attributions', methods=['GET'])
@login_required
def api_attributions_list():
    etab = get_current_etab()
    if not etab: return jsonify([])
    # Attributions en cours uniquement
    attrs = (AttributionCle.query
             .join(EmployeCle, AttributionCle.employe_id == EmployeCle.id)
             .filter(EmployeCle.etab_id == etab.id)
             .filter((AttributionCle.date_retour == '') | (AttributionCle.date_retour == None))
             .order_by(AttributionCle.date_attribution.desc())
             .all())
    return jsonify([a.to_dict() for a in attrs])


@app.route('/api/cles/attributions', methods=['POST'])
@login_required
def api_attributions_add():
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    data = request.get_json()
    emp_id = data.get('employe_id')
    cle_id = data.get('cle_id')
    if not emp_id or not cle_id: return jsonify({'error': 'employe_id et cle_id requis'}), 400

    emp = EmployeCle.query.filter_by(id=emp_id, etab_id=etab.id).first()
    cle = CleItem.query.filter_by(id=cle_id, etab_id=etab.id).first()
    if not emp or not cle: return jsonify({'error': 'Employé ou clé introuvable'}), 404

    # Vérifier disponibilité
    attribuees = AttributionCle.query.filter_by(cle_id=cle_id).filter(
        (AttributionCle.date_retour == '') | (AttributionCle.date_retour == None)
    ).count()
    if attribuees >= cle.quantite_totale:
        return jsonify({'error': f'Clé "{cle.nom}" plus disponible (toutes attribuées)'}), 409

    from datetime import date
    today = date.today().isoformat()
    a = AttributionCle(employe_id=emp_id, cle_id=cle_id,
                       date_attribution=data.get('date_attribution', today),
                       notes=data.get('notes', ''))
    db.session.add(a)
    db.session.flush()
    _log_cle_event(etab.id, 'attribution',
                   f'Clé "{cle.nom}" attribuée à {emp.prenom} {emp.nom}')
    db.session.commit()
    return jsonify(a.to_dict()), 201


@app.route('/api/cles/attributions/<int:attr_id>/retour', methods=['POST'])
@login_required
def api_attribution_retour(attr_id):
    etab = get_current_etab()
    a = AttributionCle.query.get_or_404(attr_id)
    # Vérifier appartenance
    emp = EmployeCle.query.filter_by(id=a.employe_id, etab_id=etab.id).first()
    if not emp: return jsonify({'error': 'Non autorise'}), 403
    from datetime import date
    a.date_retour = request.get_json().get('date_retour', date.today().isoformat())
    _log_cle_event(etab.id, 'retour',
                   f'Clé "{a.cle.nom if a.cle else "?"}" rendue par {emp.prenom} {emp.nom}')
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/cles/historique', methods=['GET'])
@login_required
def api_cles_historique():
    etab = get_current_etab()
    if not etab: return jsonify([])
    hist = (HistoriqueCle.query.filter_by(etab_id=etab.id)
            .order_by(HistoriqueCle.id.desc()).limit(200).all())
    return jsonify([h.to_dict() for h in hist])


# ══════════════════════════════════════════════════════
#  CONTRATS PDF PRESTATAIRES
# ══════════════════════════════════════════════════════

@app.route('/api/prestataires/contrats', methods=['GET'])
@login_required
def api_contrats_list():
    etab = get_current_etab()
    if not etab: return jsonify([])
    contrats = (ContratPDF.query.filter_by(etab_id=etab.id)
                .order_by(ContratPDF.id.desc()).all())
    return jsonify([c.to_dict() for c in contrats])


@app.route('/api/prestataires/<int:pid>/contrats', methods=['POST'])
@login_required
def api_contrat_upload(pid):
    from datetime import date as _date
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Seuls les PDF sont acceptés'}), 400
    presta = Prestataire.query.filter_by(id=pid, etab_id=etab.id).first()
    if not presta: return jsonify({'error': 'Prestataire introuvable'}), 404
    contenu = f.read()
    c = ContratPDF(
        etab_id=etab.id, presta_id=pid,
        nom_fichier=f.filename,
        date_ajout=_date.today().isoformat(),
        contenu=contenu
    )
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201


@app.route('/api/prestataires/contrats/<int:cid>', methods=['GET'])
@login_required
def api_contrat_download(cid):
    from flask import send_file
    import io
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    c = ContratPDF.query.filter_by(id=cid, etab_id=etab.id).first()
    if not c: return jsonify({'error': 'Contrat introuvable'}), 404
    return send_file(
        io.BytesIO(c.contenu),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=c.nom_fichier
    )


@app.route('/api/prestataires/contrats/<int:cid>', methods=['DELETE'])
@login_required
def api_contrat_delete(cid):
    etab = get_current_etab()
    if not etab: return jsonify({'error': 'Pas d\'etablissement'}), 400
    c = ContratPDF.query.filter_by(id=cid, etab_id=etab.id).first()
    if not c: return jsonify({'error': 'Contrat introuvable'}), 404
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
#  CHARGEMENT BDPM AU DEMARRAGE (background thread)
# ══════════════════════════════════════════════════════

def _auto_load_bdpm():
    """Charge la BDPM en arriere-plan au demarrage."""
    import time
    time.sleep(3)  # Laisser le serveur démarrer d'abord
    try:
        ok, msg = _bdpm_load()
        print(f"[BDPM] {msg}" if ok else f"[BDPM] Echec: {msg}")
    except Exception as e:
        print(f"[BDPM] Erreur auto-load: {e}")

import threading
_bdpm_thread = threading.Thread(target=_auto_load_bdpm, daemon=True)
_bdpm_thread.start()

# ══════════════════════════════════════════════════════
#  DEMARRAGE
# ══════════════════════════════════════════════════════

if __name__ == '__main__':
    app.run(debug=True, port=5000)
